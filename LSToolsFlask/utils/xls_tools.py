import os.path

import openpyxl


class XlsTools(object):
    def __init__(self, file_path, title=None):
        self.file_path = file_path
        self.workbook = None
        self.current_sheet = None
        try:
            self.workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            self.workbook = openpyxl.Workbook()
            self.workbook.save(file_path)
            self.workbook = openpyxl.load_workbook(file_path)

        self.current_sheet = self.workbook.active
        if title:
            self.write_row(title, row=1)
            self.save()

    def switch_sheet(self, sheet_name):
        """
        切换到指定名称的工作表。
        :param sheet_name: 工作表的名称。
        """
        if sheet_name in self.workbook.sheetnames:
            self.current_sheet = self.workbook[sheet_name]
        else:
            raise ValueError(f"工作表 {sheet_name} 不存在。")

    def read_cell(self, row, column):
        return self.current_sheet.cell(row=row, column=column).value

    def write_cell(self, row, column, value):
        self.current_sheet.cell(row=row, column=column).value = value

    def write_row(self, data, row=0):
        new_row = row
        if new_row == 0:
            last_row = self.current_sheet.max_row
            new_row = last_row + 1
        for idx, value in enumerate(data, start=1):
            self.write_cell(new_row, idx, value)

    def delete_row(self, row):
        self.current_sheet.delete_rows(row)

    def read_all_rows(self):
        all_rows = []
        for row in self.current_sheet.iter_rows(values_only=True):
            curr_row = []
            t = []
            for cell in row:
                if cell is None:
                    t.append(None)
                elif t:
                    t.append(cell)
                    curr_row += t
                    t = []
                else:
                    curr_row.append(cell)

            if curr_row:  # 忽略完全为空的行
                all_rows.append(tuple(curr_row))
        return all_rows

    def save(self):
        self.workbook.save(self.file_path)

    def close(self):
        self.workbook.close()
